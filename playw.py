import os
import random
from faker import Faker
from loguru import logger
from fake_useragent import FakeUserAgent
from patchright.async_api import async_playwright
# from playwright_stealth import stealth_async # Removed
import string
import secrets
import asyncio # Ensure asyncio is imported at the top
from tenacity import retry as tenacity_retry, stop_after_attempt, wait_exponential, retry_if_exception_type
from playwright.async_api import TimeoutError as PlaywrightTimeoutError
import openpyxl
from openpyxl import Workbook
from openpyxl.utils.exceptions import InvalidFileException # For checking if file is valid excel
import pyautogui
import pygetwindow

# the logger configs
info_log = logger.info
success_log = logger.success
error_log = logger.error
debug_log = logger.debug
warning_log = logger.warning

# the root name and family txt files
dir_path = os.path.dirname(os.path.abspath(__file__))
name_file = os.path.join(dir_path, "usersdata/names.txt")
family_file = os.path.join(dir_path, "usersdata/family.txt")

# Excel configuration
EXCEL_FILE_NAME = "created_outlook_accounts.xlsx"
EXCEL_HEADERS = ["Email", "Password", "First Name", "Last Name", "Birth Date"]

def save_account_to_excel(email, password, first_name, last_name, birth_date_str):
    workbook = None
    sheet = None
    file_exists_and_is_accessible = os.path.exists(EXCEL_FILE_NAME) and os.access(EXCEL_FILE_NAME, os.W_OK)

    try:
        if file_exists_and_is_accessible:
            try:
                workbook = openpyxl.load_workbook(EXCEL_FILE_NAME)
                sheet = workbook.active
            except InvalidFileException:
                info_log(f"{EXCEL_FILE_NAME} is not a valid Excel file or is corrupted. Creating a new one.")
                os.remove(EXCEL_FILE_NAME) # Attempt to remove corrupted file
                workbook = Workbook()
                sheet = workbook.active
                sheet.append(EXCEL_HEADERS)
            except Exception as e: # Other errors during load (e.g. file locked)
                error_log(f"Could not load {EXCEL_FILE_NAME}: {e}. Will try to create new if it was removed or write to a new one.")
                # If removal failed or wasn't attempted, try creating a new one.
                if not os.path.exists(EXCEL_FILE_NAME): # If it was removed or never existed
                    workbook = Workbook()
                    sheet = workbook.active
                    sheet.append(EXCEL_HEADERS)
                else: # File still exists but couldn't be loaded, avoid overwriting potentially good data with just headers
                    error_log(f"Cannot proceed with saving to {EXCEL_FILE_NAME} due to load error and file still existing.")
                    raise # Propagate the error
        else: # File does not exist or is not writable
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(EXCEL_HEADERS)

        # Ensure sheet is valid and headers are present if it's an existing sheet
        if sheet is not None and sheet.max_row > 0:
            current_headers = [sheet.cell(row=1, column=i+1).value for i in range(len(EXCEL_HEADERS))]
            if current_headers != EXCEL_HEADERS:
                # This case is tricky: headers mismatch. Could be a different file.
                # For simplicity, we'll append, but ideally, this might need more sophisticated handling or a new sheet/file.
                warning_log(f"Header mismatch in {EXCEL_FILE_NAME}. Appending data anyway.")
                if sheet.max_row == 1 and all(c is None for c in current_headers): # Empty first row
                    for i, header_text in enumerate(EXCEL_HEADERS):
                        sheet.cell(row=1, column=i + 1, value=header_text)
        elif sheet is not None and (sheet.max_row == 0 or sheet.cell(row=1, column=1).value is None): # Sheet is empty or first cell is empty
             sheet.append(EXCEL_HEADERS)


        if sheet is None: # If sheet could not be initialized
            error_log(f"Failed to obtain a valid sheet in {EXCEL_FILE_NAME}. Cannot save account.")
            raise # Propagate the error

        sheet.append([email, password, first_name, last_name, birth_date_str])
        workbook.save(EXCEL_FILE_NAME)
        success_log(f"Successfully saved account {email} to {EXCEL_FILE_NAME}")

    except PermissionError:
        error_log(f"Permission denied when trying to save {EXCEL_FILE_NAME}. Make sure the file is not open in Excel or another program and you have write permissions.")
        raise # Propagate the error
    except Exception as e:
        error_log(f"An unexpected error occurred while saving to Excel: {e}")
        raise # Propagate the error


# function to get the random name and family
def get_random_data():
    with open(name_file, "r") as f:
        names = f.readlines()
    first_name = random.choice(names).strip()
    with open(family_file, "r") as f:
        families = f.readlines()
    last_name = random.choice(families).strip()
    birth_date = Faker().date_of_birth(minimum_age=18, maximum_age=80)
    email_username = f"{first_name}{last_name}{random.randint(1, 999999)}"
    return {"first_name": first_name, "last_name": last_name, "email_username": email_username, "birth_date": birth_date}

def generate_strong_password(length=12):
    if length < 4: length = 12 # Ensure min length for variety, default to 12
    char_sets = [string.ascii_lowercase, string.ascii_uppercase, string.digits, string.punctuation]
    # Start password with one character from each set
    password_list = [secrets.choice(s) for s in char_sets]
    # Characters for the rest of the password
    all_chars = "".join(char_sets)
    # Fill remaining length if length > 4
    if length > len(password_list):
        password_list.extend(secrets.choice(all_chars) for _ in range(length - len(password_list)))
    elif length < len(password_list): # Should not happen if length >=4
        password_list = password_list[:length]
    random.shuffle(password_list)
    return "".join(password_list)

# Default retry settings for tenacity
RETRY_WAIT = wait_exponential(multiplier=1, min=1, max=5) # Wait 1s, then 2s, then 4s (max 5s for any single wait)
RETRY_STOP = stop_after_attempt(3) # Try 3 times in total
RETRY_ON_EXCEPTION = retry_if_exception_type(PlaywrightTimeoutError)

async def os_mouse_to_element_center(page, locator, duration_seconds=0.5):
    """Moves the OS mouse cursor to the center of the given Playwright locator."""
    info_log(f"DEBUG_OS_MOUSE: Attempting OS-level mouse move to center of locator: {locator}")
    try:
        bounding_box = await locator.bounding_box()
        if not bounding_box:
            warning_log(f"DEBUG_OS_MOUSE: Cannot get bounding_box for OS mouse move for locator: {locator}. Skipping OS move.")
            return False
        info_log(f"DEBUG_OS_MOUSE: Bounding box found: {bounding_box}")

        element_center_x_viewport = bounding_box['x'] + bounding_box['width'] / 2
        element_center_y_viewport = bounding_box['y'] + bounding_box['height'] / 2
        info_log(f"DEBUG_OS_MOUSE: Element viewport center: X={element_center_x_viewport:.0f}, Y={element_center_y_viewport:.0f}")

        page_title = await page.title()
        info_log(f"DEBUG_OS_MOUSE: Page title from Playwright: '{page_title}'")

        browser_window = None
        if not page_title:
            info_log("DEBUG_OS_MOUSE: Page title is empty. Attempting fallback to find browser window.")
            all_windows = pygetwindow.getAllWindows()
            info_log(f"DEBUG_OS_MOUSE: Fallback - Found {len(all_windows)} total windows.")
            potential_browsers = [w for w in all_windows if w.visible and w.width > 300 and w.height > 300]
            info_log(f"DEBUG_OS_MOUSE: Fallback - Found {len(potential_browsers)} potential browser windows.")
            if not potential_browsers:
                warning_log("DEBUG_OS_MOUSE: Fallback - Could not find any suitable potential browser window. Skipping OS move.")
                return False
            browser_window = potential_browsers[0] 
            info_log(f"DEBUG_OS_MOUSE: Fallback - Selected browser window: '{browser_window.title}' L:{browser_window.left}, T:{browser_window.top}, W:{browser_window.width}, H:{browser_window.height}")
        else:
            windows = pygetwindow.getWindowsWithTitle(page_title)
            info_log(f"DEBUG_OS_MOUSE: Found {len(windows)} windows matching title '{page_title}': {windows}")
            if windows:
                for i, win in enumerate(windows):
                    # Safely get attributes for logging, defaulting to False or None if not present
                    is_active_log = getattr(win, 'isActive', False)
                    is_maximized_log = getattr(win, 'isMaximized', False)
                    is_visible_log = getattr(win, 'visible', False) # Changed to win.visible
                    win_left_log = getattr(win, 'left', 'N/A')
                    win_top_log = getattr(win, 'top', 'N/A')
                    info_log(f"DEBUG_OS_MOUSE: Window {i} - Title: '{win.title}', Active: {is_active_log}, Maximized: {is_maximized_log}, Visible: {is_visible_log}, L:{win_left_log}, T:{win_top_log}")
                    
                    # Use getattr for conditions too, to be safe, and use win.visible
                    is_active = getattr(win, 'isActive', False)
                    is_maximized = getattr(win, 'isMaximized', False)
                    is_actually_visible = getattr(win, 'visible', False) # Changed to win.visible

                    if is_actually_visible and (is_active or is_maximized): # Prioritize visible and active/maximized
                        if browser_window is None: 
                           browser_window = win
                           info_log(f"DEBUG_OS_MOUSE: Selected window (priority - active/maximized & visible): '{win.title}' L:{getattr(win, 'left', 'N/A')}, T:{getattr(win, 'top', 'N/A')}")
                
                # Fallback if no active/maximized visible window was found, but some windows matching title exist
                if not browser_window and windows:
                    # Try to find any simply visible window from the list first
                    for i, win in enumerate(windows):
                        if getattr(win, 'visible', False):
                            browser_window = win
                            info_log(f"DEBUG_OS_MOUSE: Selected window (fallback - first visible): '{win.title}' L:{getattr(win, 'left', 'N/A')}, T:{getattr(win, 'top', 'N/A')}")
                            break
                    # If still no browser_window (e.g. none were visible), take the first one from the original list as a last resort
                    if not browser_window:
                        browser_window = windows[0]
                        info_log(f"DEBUG_OS_MOUSE: Selected window (fallback - truly first in list, visibility uncertain): '{browser_window.title}' L:{getattr(browser_window, 'left', 'N/A')}, T:{getattr(browser_window, 'top', 'N/A')}")
        
        if not browser_window:
            warning_log(f"DEBUG_OS_MOUSE: Could not find/select browser window (title: '{page_title}'). Skipping OS move.")
            return False
        
        info_log(f"DEBUG_OS_MOUSE: Final selected browser window: Title: '{browser_window.title}', L:{browser_window.left}, T:{browser_window.top}, W:{browser_window.width}, H:{browser_window.height}")

        window_left, window_top = browser_window.left, browser_window.top

        browser_chrome_height_estimate = 80 
        info_log(f"DEBUG_OS_MOUSE: Using browser chrome height estimate: {browser_chrome_height_estimate}")

        screen_x = window_left + element_center_x_viewport
        screen_y = window_top + browser_chrome_height_estimate + element_center_y_viewport
        info_log(f"DEBUG_OS_MOUSE: Calculated OS screen coords: X={screen_x:.0f}, Y={screen_y:.0f}")
        
        # For safety during debugging, let's check if coords are wildly off (e.g., negative)
        if screen_x < 0 or screen_y < 0:
            warning_log(f"DEBUG_OS_MOUSE: Calculated screen coordinates are negative ({screen_x}, {screen_y}). This is likely an error. Skipping OS move.")
            return False
        
        # Also check against screen size if possible
        try:
            s_width, s_height = pyautogui.size()
            info_log(f"DEBUG_OS_MOUSE: Screen size from PyAutoGUI: Width={s_width}, Height={s_height}")
            if screen_x >= s_width or screen_y >= s_height:
                warning_log(f"DEBUG_OS_MOUSE: Calculated screen coordinates ({screen_x}, {screen_y}) are outside screen bounds ({s_width}, {s_height}). Skipping OS move.")
                return False
        except Exception as e_screen_size:
            warning_log(f"DEBUG_OS_MOUSE: Could not get screen size from PyAutoGUI: {e_screen_size}. Proceeding with move cautiously.")


        info_log(f"DEBUG_OS_MOUSE: Calling pyautogui.moveTo({screen_x:.0f}, {screen_y:.0f}, duration={duration_seconds})")
        pyautogui.moveTo(screen_x, screen_y, duration=duration_seconds)
        success_log(f"DEBUG_OS_MOUSE: OS mouse move call completed for locator: {locator}. Current OS mouse pos: {pyautogui.position()}")
        return True

    except NotImplementedError as e:
        # This can happen on some OS/environments where pyautogui can't get screen info or control mouse
        error_log(f"PyAutoGUI NotImplementedError during OS-level mouse move: {e}. This feature may not work on your system.")
        return False
    except Exception as e:
        error_log(f"Error during OS-level mouse move for locator {locator}: {e}")
        return False

# Helper functions with tenacity
@tenacity_retry(wait=RETRY_WAIT, stop=RETRY_STOP, retry=RETRY_ON_EXCEPTION)
async def robust_goto(page, url, **kwargs):
    info_log(f"Attempting to navigate to {url} with retries...")
    await page.goto(url, **kwargs)
    success_log(f"Successfully navigated to {url}")

@tenacity_retry(wait=RETRY_WAIT, stop=RETRY_STOP, retry=RETRY_ON_EXCEPTION)
async def robust_wait_for_selector(page, selector, **kwargs):
    info_log(f"Attempting to wait for selector '{selector}' with retries...")
    element = await page.wait_for_selector(selector, **kwargs)
    success_log(f"Successfully found selector '{selector}'")
    return element

@tenacity_retry(wait=RETRY_WAIT, stop=RETRY_STOP, retry=RETRY_ON_EXCEPTION)
async def robust_hover(page, locator, use_os_mouse_move=False, os_mouse_duration=0.5, **kwargs):
    info_log(f"Attempting to hover locator '{locator}' with human-like steps...")

    if use_os_mouse_move:
        info_log(f"Pre-hover: Initiating OS-level mouse move for locator '{locator}'")
        await os_mouse_to_element_center(page, locator, duration_seconds=os_mouse_duration)
        # Optional: Short delay after OS mouse movement for stability before Playwright takes over
        await asyncio.sleep(random.uniform(0.1, 0.3))

    # Proceed with Playwright's hover logic
    try:
        await locator.scroll_into_view_if_needed(timeout=5000) # Scroll into view first
    except Exception as e:
        warning_log(f"Failed to scroll locator '{locator}' into view for Playwright hover: {e}. Proceeding with hover attempt.")

    bounding_box = await locator.bounding_box()
    if bounding_box:
        target_x = bounding_box['x'] + bounding_box['width'] / 2
        target_y = bounding_box['y'] + bounding_box['height'] / 2
        
        num_steps = kwargs.get('steps', 5) 
        info_log(f"Playwright mouse moving to ({target_x:.0f}, {target_y:.0f}) in {num_steps} steps for locator '{locator}'.")
        await page.mouse.move(target_x, target_y, steps=num_steps)
        
        success_log(f"Successfully hovered (Playwright) over locator '{locator}'")
    else:
        warning_log(f"Could not get bounding_box for Playwright hover on locator '{locator}'. Falling back to locator.hover().")
        try:
            await locator.scroll_into_view_if_needed(timeout=5000)
        except Exception as e:
            warning_log(f"Failed to scroll locator '{locator}' into view for fallback Playwright hover: {e}")
        await locator.hover(**{k: v for k, v in kwargs.items() if k not in ('use_os_mouse_move', 'os_mouse_duration')}) # fallback
        success_log(f"Successfully hovered (Playwright fallback) over locator '{locator}'")

@tenacity_retry(wait=RETRY_WAIT, stop=RETRY_STOP, retry=RETRY_ON_EXCEPTION)
async def robust_click(locator, use_os_mouse_move=False, os_mouse_duration=0.5, **kwargs):
    page = locator.page # Get the page object from the locator
    info_log(f"Attempting to human-like hover and click locator '{locator}' (OS mouse: {use_os_mouse_move})...")

    # Perform a Playwright-controlled human-like hover first.
    # This hover will internally handle the OS mouse move if use_os_mouse_move is True.
    await robust_hover(page, locator, use_os_mouse_move=use_os_mouse_move, os_mouse_duration=os_mouse_duration, steps=kwargs.get('hover_steps', 5))
    
    # Short delay between hover and click, can be randomized
    await asyncio.sleep(random.uniform(0.1, 0.3)) 

    # Proceed with the click
    info_log(f"Performing Playwright click on locator '{locator}'")
    try:
        # Ensure it's still in view; robust_hover should have handled scrolling, but a quick check is okay.
        await locator.scroll_into_view_if_needed(timeout=2000) 
    except Exception as e:
        warning_log(f"Failed to scroll locator '{locator}' into view just before click (should be minor): {e}. Proceeding with click attempt.")
    
    # Pass original kwargs except those handled by robust_hover or specific to robust_click's OS move logic
    click_kwargs = {k: v for k, v in kwargs.items() if k not in ('hover_steps', 'use_os_mouse_move', 'os_mouse_duration')}
    await locator.click(**click_kwargs) 
    success_log(f"Successfully clicked (Playwright) locator '{locator}'")

@tenacity_retry(wait=RETRY_WAIT, stop=RETRY_STOP, retry=RETRY_ON_EXCEPTION)
async def robust_type(locator, text, **kwargs):
    original_timeout = kwargs.get('timeout')
    original_delay = kwargs.get('delay')

    log_message = f"Attempting to type text '{text[:20]}...' into locator '{locator}' char-by-char with random delays."
    if original_delay is not None:
        # Inform that the 'delay' kwarg is now handled differently
        log_message += f" (Note: 'delay={original_delay}' for locator.type is superseded by internal per-char random delays)."
    if original_timeout is not None:
        # Inform that 'timeout' for the whole operation is now mainly governed by tenacity
        log_message += f" (Note: Overall operation timeout is primarily handled by tenacity retries, not 'timeout={original_timeout}' for a single type op)."
    info_log(log_message)

    try:
        await locator.scroll_into_view_if_needed(timeout=5000) # Standard timeout for scrolling
    except Exception as e:
        warning_log(f"Failed to scroll locator '{locator}' into view for type: {e}. Proceeding with type attempt.")

    try:
        # Click to focus the element before typing. Use a reasonable timeout for the click itself.
        await locator.click(timeout=kwargs.get('click_timeout', 5000)) # Allow override or use default
    except Exception as e:
        warning_log(f"Failed to click/focus locator '{locator}' before typing: {e}. Attempting to type anyway.")

    for char_index, char_to_press in enumerate(text):
        try:
            await locator.press(char_to_press)
            # Randomized delay after each character press (50ms to 250ms)
            char_delay = random.uniform(0.05, 0.25)
            await asyncio.sleep(char_delay)
        except PlaywrightTimeoutError as e:
            error_log(f"Playwright TimeoutError while pressing character '{char_to_press}' (index {char_index}) for locator '{locator}': {e}")
            raise # Re-raise to trigger tenacity retry for the whole robust_type
        except Exception as e:
            error_log(f"Error pressing character '{char_to_press}' (index {char_index}) for locator '{locator}': {e}")
            raise # Re-raise to trigger tenacity

    success_log(f"Successfully typed text into locator '{locator}' using char-by-char method.")

@tenacity_retry(wait=RETRY_WAIT, stop=RETRY_STOP, retry=RETRY_ON_EXCEPTION)
async def robust_select_option(locator, value, **kwargs):
    info_log(f"Attempting to select option '{value}' for locator '{locator}' with retries...")
    try:
        await locator.scroll_into_view_if_needed(timeout=5000)
    except Exception as e:
        warning_log(f"Failed to scroll locator '{locator}' into view for select_option: {e}. Proceeding with select_option attempt.")
    await locator.select_option(value=value, **kwargs)
    success_log(f"Successfully selected option '{value}' for locator '{locator}'")

@tenacity_retry(wait=RETRY_WAIT, stop=RETRY_STOP, retry=RETRY_ON_EXCEPTION)
async def robust_wait_for_load_state(page, state="domcontentloaded", **kwargs):
    info_log(f"Attempting to wait for load state '{state}' with retries...")
    await page.wait_for_load_state(state, **kwargs)
    success_log(f"Successfully reached load state '{state}'")

# function to get the random date of birt and region

# this code will open the browser and go to the signup page
async def main():
    async with async_playwright() as p:
        info_log("Starting the browser")
        # this code will open the browser and go to the signup page
        browser = await p.chromium.launch(headless=False) # Removed args=launch_args to use patchright defaults

        # Define common viewport sizes
        common_viewports = [{'width': 1920, 'height': 1080}, {'width': 1366, 'height': 768}, {'width': 1280, 'height': 720}, {'width': 1600, 'height': 900}, {'width': 1440, 'height': 900}, {'width': 1280, 'height': 800}, {'width': 1280, 'height': 1024},{'width': 1024, 'height': 768}]
        # Select a random viewport
        random_viewport = random.choice(common_viewports)
        info_log(f"Setting random viewport: {random_viewport['width']}x{random_viewport['height']}")

        context = await browser.new_context(user_agent=FakeUserAgent().random, viewport=random_viewport)
        # await context.add_init_script(init_script) # Ensure it runs on every new document

        page = await context.new_page()
        # await stealth_async(page) # Removed

        # Intercept network requests to block images, media, fonts, and stylesheets
        async def handle_route(route):
            if route.request.resource_type in ["image", "media"]: # Allow fonts and stylesheets
                await route.abort()
            else:
                await route.continue_()
        await page.route("**/*", handle_route)

        try: # this code will try to go to the signup page
            await robust_goto(page, "https://signup.live.com/signup")
        except Exception as e:
            error_log(f"Error navigating to signup page: {e}")
            raise # Propagate the error
        # input the email to the first page
        random_data = get_random_data()
        email_input_selector = 'input[aria-label="Email"]'
        email_input = page.locator(email_input_selector)
        info_log(f"Attempting to find email input with selector: {email_input_selector}")
        # this code will try to input the email to the first page
        try:
            # Best practice: Hover to simulate mouse movement
            info_log("Hovering over the email input.")
            await robust_hover(page, email_input, use_os_mouse_move=True, os_mouse_duration=0.7, timeout=5000) # Wait up to 5s for hover
            # Best practice: Click to focus (fill also does this, but explicit click is fine)
            info_log("Clicking the email input.")
            await asyncio.sleep(random.uniform(1.0, 3.0))
            await robust_click(email_input, use_os_mouse_move=True, os_mouse_duration=0.7, timeout=5000) # Wait up to 5s for click
            # Type the email into the input field, simulating human behavior
            info_log(f"Attempting to type email: {random_data['email_username']}") # Updated log message
            await asyncio.sleep(random.uniform(1.0, 3.0))
            await robust_type(email_input, f"{random_data['email_username']}@outlook.com", delay=random.uniform(90, 300), timeout=15000)
            success_log("Successfully typed the email input.") # Updated log message
            # Click the next button
            next_button_selector = "[data-testid='primaryButton']"
            next_button = page.locator(next_button_selector)
            info_log(f"Attempting to find 'Next' button with selector: {next_button_selector}")
            # Hover to simulate human-like mouse movement
            info_log("Hovering over the 'Next' button.")
            await robust_hover(page, next_button, use_os_mouse_move=True, os_mouse_duration=0.7, timeout=5000)  # Wait up to 5 seconds for hover
            # Click the 'Next' button
            info_log("Clicking the 'Next' button.")
            await asyncio.sleep(random.uniform(1.0, 3.0))
            await robust_click(next_button, use_os_mouse_move=True, os_mouse_duration=0.7, timeout=5000)  # Wait up to 5 seconds for click
            success_log("Successfully clicked the 'Next' button.")
            
            info_log("Waiting for the next page to load (password page expected).")
            try:
                await robust_wait_for_load_state(page, 'domcontentloaded', timeout=15000) # Wait for DOM content
                success_log("Next page DOM content loaded.")
            except Exception as e:
                error_log(f"Error waiting for page load state after clicking next: {e}")
                # Potentially add a screenshot here for debugging if it fails often
                # await page.screenshot(path="error_screenshot_pageload.png")
                raise # Propagate the error

        except Exception as e:
            error_log(f"Error interacting with email input: {e}")
            raise # Propagate the error
        # this code will try to insert the password to the second page
        try:
            info_log("Waiting for the password input to appear")
            # Wait for the password input field to be present and visible
            password_selector = 'input[type="password"][autocomplete="new-password"]'
            await robust_wait_for_selector(page, password_selector, timeout=20000) # Wait up to 20s
            password_input_locator = page.locator(password_selector)
            password_to_save = generate_strong_password() # Assign the generated password
            info_log("Hovering over the password input.")
            await robust_hover(page, password_input_locator, use_os_mouse_move=True, os_mouse_duration=0.7, timeout=5000)
            info_log("Clicking the password input.")
            await asyncio.sleep(random.uniform(1.0, 3.0))
            await robust_click(password_input_locator, use_os_mouse_move=True, os_mouse_duration=0.7, timeout=5000)
            info_log(f"Attempting to type password...") # Removed password from log for security
            await asyncio.sleep(random.uniform(1.0, 3.0))
            await robust_type(password_input_locator, text=password_to_save, delay=random.uniform(90, 300), timeout=15000) # Use password_to_save
            next_button_password_selector = "[data-testid='primaryButton']"
            next_button_password = page.locator(next_button_password_selector)
            await asyncio.sleep(random.uniform(1.0, 3.0))
            await robust_hover(page, next_button_password, use_os_mouse_move=True, os_mouse_duration=0.7, timeout=5000)
            await asyncio.sleep(random.uniform(1.0, 3.0))
            await robust_click(next_button_password, use_os_mouse_move=True, os_mouse_duration=0.7, timeout=5000)
            success_log("Successfully clicked the 'Next' button after password input.")
            info_log("Waiting for the birth date page to load.")
            try:
                await robust_wait_for_load_state(page, 'domcontentloaded', timeout=15000) # Wait for DOM content
                success_log("Birth date page DOM content loaded.")
            except Exception as e:
                error_log(f"Error waiting for birth date page load state: {e}")
                raise # Propagate the error

        except Exception as e:
            error_log(f"Error interacting with password input: {e}")
            raise # Propagate the error
        # give the mont, day, year and the region to the third page
        try:
            info_log("Waiting for the birth month dropdown to appear.")
            month_selector = 'select[name="BirthMonth"]'
            await robust_wait_for_selector(page, month_selector, timeout=10000) # Wait up to 10s
            
            birth_date_obj = random_data['birth_date']
            birth_month_value = str(birth_date_obj.month)
            info_log(f"Attempting to select birth month: {birth_month_value}")
            month_dropdown_locator = page.locator(month_selector)
            await asyncio.sleep(random.uniform(1.0, 3.0))
            await robust_hover(page, month_dropdown_locator, use_os_mouse_move=True, os_mouse_duration=0.7, timeout=5000)
            await asyncio.sleep(random.uniform(1.0, 3.0))
            await robust_select_option(month_dropdown_locator, value=birth_month_value)
            success_log(f"Successfully selected birth month: {birth_month_value}")

            # Select Day
            info_log("Waiting for the birth day dropdown to appear.")
            day_selector = 'select[name="BirthDay"]'
            await robust_wait_for_selector(page, day_selector, timeout=10000)
            day_dropdown_locator = page.locator(day_selector)
            birth_day_value = str(birth_date_obj.day) # birth_date_obj is already defined
            info_log(f"Attempting to select birth day: {birth_day_value}")
            await asyncio.sleep(random.uniform(1.0, 3.0))
            await robust_hover(page, day_dropdown_locator, use_os_mouse_move=True, os_mouse_duration=0.7, timeout=5000)
            await asyncio.sleep(random.uniform(1.0, 3.0))
            await robust_select_option(day_dropdown_locator, value=birth_day_value)
            success_log(f"Successfully selected birth day: {birth_day_value}")
            
            # Input Year
            info_log("Waiting for the birth year input to appear.")
            year_selector = 'input[name="BirthYear"]'
            await robust_wait_for_selector(page, year_selector, timeout=10000)
            year_input_locator = page.locator(year_selector)
            birth_year_value = str(birth_date_obj.year) # birth_date_obj is already defined
            info_log(f"Attempting to input birth year: {birth_year_value}")
            await asyncio.sleep(random.uniform(1.0, 3.0))
            await robust_hover(page, year_input_locator, use_os_mouse_move=True, os_mouse_duration=0.7, timeout=5000)
            await asyncio.sleep(random.uniform(1.0, 3.0))
            await robust_click(year_input_locator, use_os_mouse_move=True, os_mouse_duration=0.7, timeout=5000)
            await asyncio.sleep(random.uniform(1.0, 3.0))
            await robust_type(year_input_locator, text=birth_year_value, delay=random.uniform(90, 250))
            success_log(f"Successfully inputted birth year: {birth_year_value}")
            # Click the Next button after birth date input
            info_log("Attempting to find the 'Next' button after birth date input.")
            next_button_birthdate_selector = "[data-testid='primaryButton']"
            next_button_birthdate = page.locator(next_button_birthdate_selector)
            info_log(f"Found 'Next' button with selector: {next_button_birthdate_selector}")
            info_log("Hovering over the 'Next' button (after birth date).")
            await robust_hover(page, next_button_birthdate, use_os_mouse_move=True, os_mouse_duration=0.7, timeout=5000)
            info_log("Clicking the 'Next' button (after birth date).")
            await asyncio.sleep(random.uniform(1.0, 3.0))
            await robust_click(next_button_birthdate, use_os_mouse_move=True, os_mouse_duration=0.7, timeout=5000)
            success_log("Successfully clicked the 'Next' button after birth date input.")

            info_log("Waiting for the name input page to load.")
            try:
                await robust_wait_for_load_state(page, 'domcontentloaded', timeout=15000)
                success_log("Name input page DOM content loaded.")
            except Exception as e:
                error_log(f"Error waiting for name input page load state: {e}")
                raise # Propagate the error

        except Exception as e:
            error_log(f"Error interacting with birth date input: {e}")
            raise # Propagate the error
        # Input first name and last name for the next page
        try:
            info_log("Waiting for first name input to appear.")
            first_name_selector = "input[id='firstNameInput']"
            await robust_wait_for_selector(page, first_name_selector, timeout=10000)
            first_name_input_locator = page.locator(first_name_selector)
            user_first_name = random_data['first_name']
            info_log(f"Attempting to input first name: {user_first_name}")
            await asyncio.sleep(random.uniform(1.0, 3.0))
            await robust_hover(page, first_name_input_locator, use_os_mouse_move=True, os_mouse_duration=0.7, timeout=5000)
            await asyncio.sleep(random.uniform(1.0, 3.0))
            await robust_click(first_name_input_locator, use_os_mouse_move=True, os_mouse_duration=0.7, timeout=5000)
            await asyncio.sleep(random.uniform(1.0, 3.0))
            await robust_type(first_name_input_locator, text=user_first_name, delay=random.uniform(90, 250))
            success_log(f"Successfully inputted first name: {user_first_name}")

            info_log("Waiting for last name input to appear.")
            last_name_selector = "input[id='lastNameInput']"
            await robust_wait_for_selector(page, last_name_selector, timeout=10000)
            last_name_input_locator = page.locator(last_name_selector)
            user_last_name = random_data['last_name']
            info_log(f"Attempting to input last name: {user_last_name}")
            await asyncio.sleep(random.uniform(1.0, 3.0))
            await robust_hover(page, last_name_input_locator, use_os_mouse_move=True, os_mouse_duration=0.7, timeout=5000)
            await asyncio.sleep(random.uniform(1.0, 3.0))
            await robust_click(last_name_input_locator, use_os_mouse_move=True, os_mouse_duration=0.7, timeout=5000)
            await asyncio.sleep(random.uniform(1.0, 3.0))
            await robust_type(last_name_input_locator, text=user_last_name, delay=random.uniform(90, 250))
            success_log(f"Successfully inputted last name: {user_last_name}")
            # Click the Next button after name input
            info_log("Attempting to find the 'Next' button after name input.")
            next_button_name_selector = "[data-testid='primaryButton']"
            next_button_name = page.locator(next_button_name_selector)
            info_log(f"Found 'Next' button with selector: {next_button_name_selector}")
            info_log("Hovering over the 'Next' button (after name input).")
            await robust_hover(page, next_button_name, use_os_mouse_move=True, os_mouse_duration=0.7, timeout=5000)
            info_log("Clicking the 'Next' button (after name input).")
            await asyncio.sleep(random.uniform(1.0, 3.0))
            await robust_click(next_button_name, use_os_mouse_move=True, os_mouse_duration=0.7, timeout=5000)
            success_log("Successfully clicked the 'Next' button after name input.")
            
            try:
                # --- Enhanced CAPTCHA Detection ---
                info_log("Checking for CAPTCHA page...")
                # Give the page a moment to settle before starting checks
                await page.wait_for_timeout(3000) 

                is_captcha_visible = False
                captcha_detection_reason = "No specific CAPTCHA element detected by any method."

                # Check 1: "Solve puzzle" button
                if not is_captcha_visible:
                    try:
                        solve_puzzle_button = page.get_by_role("button", name="Solve puzzle")
                        if await solve_puzzle_button.is_visible(timeout=1000):
                            is_captcha_visible = True
                            captcha_detection_reason = "'Solve puzzle' button is visible."
                    except PlaywrightTimeoutError:
                        info_log("Timeout checking for 'Solve puzzle' button.")
                    except Exception as e:
                        warning_log(f"Error checking for 'Solve puzzle' button: {e}")
                
                # Check 2: Specific Arkose Labs iframe (data-testid='enforcementFrame')
                if not is_captcha_visible:
                    try:
                        arkose_iframe = page.locator("iframe[data-testid='enforcementFrame']")
                        if await arkose_iframe.is_visible(timeout=1000):
                            is_captcha_visible = True
                            captcha_detection_reason = "Arkose Labs iframe (enforcementFrame) is visible."
                        elif await arkose_iframe.count() > 0: # Check if it exists in DOM even if not visible
                            is_captcha_visible = True # Assuming its presence means a CAPTCHA
                            captcha_detection_reason = "Arkose Labs iframe (enforcementFrame) found in DOM (may not be immediately visible)."
                    except PlaywrightTimeoutError:
                        info_log("Timeout checking for Arkose Labs enforcementFrame.")
                    except Exception as e:
                        warning_log(f"Error checking for Arkose Labs enforcementFrame: {e}")

                # Check 3: "Let's prove you're human" heading (h1)
                if not is_captcha_visible:
                    try:
                        human_proof_heading = page.locator("h1:has-text('Let\'s prove you\'re human')")
                        if await human_proof_heading.is_visible(timeout=1000):
                            is_captcha_visible = True
                            captcha_detection_reason = "'Let\'s prove you\'re human' (h1) heading is visible."
                    except PlaywrightTimeoutError:
                        info_log("Timeout checking for 'Let\'s prove you\'re human' heading.")
                    except Exception as e:
                        warning_log(f"Error checking for 'Let\'s prove you\'re human' heading: {e}")
                
                # Check 4: Other common CAPTCHA-related headings (e.g., h2)
                if not is_captcha_visible:
                    try:
                        # Example: Look for "Please solve this puzzle" or similar text in an h2
                        # This can be expanded with other common phrases
                        puzzle_heading_h2 = page.locator("h2:has-text('Please solve this puzzle'i)") # Case-insensitive
                        if await puzzle_heading_h2.is_visible(timeout=1000):
                            is_captcha_visible = True
                            captcha_detection_reason = "Common puzzle/verification heading (h2) is visible."
                    except PlaywrightTimeoutError:
                        info_log("Timeout checking for common puzzle/verification heading (h2).")
                    except Exception as e:
                        warning_log(f"Error checking for common puzzle/verification heading (h2): {e}")

                # Check 5: Generic CAPTCHA iframes by title keywords or src attributes
                if not is_captcha_visible:
                    captcha_iframe_selectors = ["iframe[title*='challenge'i]", "iframe[title*='verification'i]", "iframe[title*='captcha'i]", "iframe[title*='security'i]", "iframe[title*='puzzle'i]", "iframe[src*='hcaptcha.com'i]", "iframe[src*='recaptcha.net'i]", "iframe[src*='google.com/recaptcha'i]", "iframe[src*='arkoselabs.com'i]", "iframe[src*='funcaptcha.com'i]", "iframe[src*='hsprotect.net']"]
                    try:
                        for selector in captcha_iframe_selectors:
                            matching_iframes = page.locator(selector)
                            count = await matching_iframes.count()
                            for i in range(count):
                                iframe_candidate = matching_iframes.nth(i)
                                # Use a very short timeout here as we are iterating
                                if await iframe_candidate.is_visible(timeout=500): 
                                    is_captcha_visible = True
                                    captcha_detection_reason = f"Generic CAPTCHA iframe visible (selector: {selector}, instance {i+1})."
                                    break # Break from inner loop (checking instances of current selector)
                            if is_captcha_visible:
                                break # Break from outer loop (checking different selectors)
                    except PlaywrightTimeoutError:
                        info_log("Timeout during generic CAPTCHA iframe checks.")
                    except Exception as e:
                        warning_log(f"Error during generic CAPTCHA iframe checks: {e}")
                
                if is_captcha_visible:
                    info_log(f"CAPTCHA DETECTED: {captcha_detection_reason}")
                else:
                    info_log("CAPTCHA NOT DETECTED based on current checks.")
            
            except Exception as e_captcha_outer:
                # This catches errors not caught by inner try-except blocks, or if initial page.wait_for_timeout fails
                error_log(f"An unexpected error occurred in the outer CAPTCHA detection block: {e_captcha_outer}")
                raise

            if is_captcha_visible:
                info_log("CAPTCHA page is loaded.")
                info_log("SCRIPT PAUSED: Please solve the CAPTCHA now in the browser window.")
                info_log("The script will wait indefinitely for the page to navigate after you solve the CAPTCHA...")
                # Wait indefinitely for navigation to complete after manual CAPTCHA solving.
                try:
                    await page.wait_for_load_state('load', timeout=0) # timeout=0 means indefinite
                    success_log("CAPTCHA appears to be solved (navigation detected). Waiting for notice page content.") # Updated log

                    # Wait for the "OK" button on the account notice page to become visible
                    notice_page_ok_button = page.get_by_role("button", name="OK")
                    info_log("Waiting for the 'OK' button on the account notice page to become visible (up to 60s).")
                    try:
                        await notice_page_ok_button.wait_for(state="visible", timeout=120000) # Wait up to 60 seconds
                        success_log("'OK' button on notice page is visible.")
                    except PlaywrightTimeoutError:
                        error_log("Timeout (60s) waiting for 'OK' button on the notice page after CAPTCHA. Page may not have loaded as expected.")
                        raise # Propagate the error
                    success_log("Proceeding to check current page after waiting for potential notice page content.") # Updated log
                except PlaywrightTimeoutError: 
                    error_log("Timeout unexpectedly occurred while waiting for navigation after manual CAPTCHA (this shouldn't happen with timeout=0).")
                    raise # Propagate the error 
                except Exception as e_nav:
                    error_log(f"An error occurred while waiting for navigation or notice page content after manual CAPTCHA: {e_nav}")
                    raise # Propagate the error
            else:
                info_log("CAPTCHA page not detected. Proceeding as if no CAPTCHA was presented.")
                

            # --- Check for privacy notice page (or other outcomes) AFTER manual CAPTCHA or if no CAPTCHA ---
            current_url = page.url
            page_title = await page.title()
            info_log(f"Landed on URL: {current_url} - Title: {page_title} (after CAPTCHA/name input)")

            if "privacynotice.account.microsoft.com" in current_url:
                info_log("Microsoft account notice page detected. Preparing to save details and click OK.")
                if password_to_save and random_data:
                    email_full = f"{random_data['email_username']}@outlook.com"
                    birth_date_str = random_data['birth_date'].strftime('%Y-%m-%d')
                    save_account_to_excel(email_full, password_to_save, random_data['first_name'], random_data['last_name'], birth_date_str)
                else:
                    error_log("Missing password or random_data, cannot save account details to Excel.")
                    raise ValueError("Missing critical data for account saving after CAPTCHA.") # Propagate error

                ok_button_locator = page.get_by_role("button", name="OK")
                # Fallback selector if get_by_role doesn't work: page.locator("button:has-text('OK')")
                if await ok_button_locator.is_visible(timeout=10000):
                    info_log("Clicking 'OK' on the Microsoft account notice page.")
                    await robust_hover(page, ok_button_locator, use_os_mouse_move=True, os_mouse_duration=0.7, timeout=5000) # Hover before click
                    await robust_click(ok_button_locator, use_os_mouse_move=True, os_mouse_duration=0.7)
                    # Wait for page to potentially change/settle after clicking OK
                    try:
                        await robust_wait_for_load_state(page, 'domcontentloaded', timeout=10000)
                        success_log("Successfully clicked 'OK' and notice page processed.")
                    except PlaywrightTimeoutError:
                        info_log("Page did not fully reload or change after clicking OK on notice, but proceeding.")
                        raise
                    except Exception as e_load:
                        warning_log(f"Error waiting for load state after clicking OK on notice: {e_load}. Proceeding.")
                        raise
                else:
                    warning_log("'OK' button not found on the notice page. Trying to proceed.")
                    raise
            
            elif "account.microsoft.com" in current_url and "verify" in current_url.lower():
                 warning_log(f"Landed on a verification page unexpectedly: {current_url}. This might indicate an unsolved CAPTCHA or other challenge.")
                 raise

            else:
                warning_log(f"Did not land on the expected Microsoft account notice page. Current URL: {current_url}, Title: {page_title}")
                raise

        except Exception as e:
            error_log(f"Error interacting with name and family input OR on the subsequent notice/final page: {e}")
            raise # Propagate the error
        finally:
            info_log("Account creation attempt finished for this iteration. Preparing to close browser resources.")
            if 'page' in locals() and page and not page.is_closed():
                info_log("Waiting for 3 seconds before closing page.")
                await page.wait_for_timeout(3000) # Brief pause before closing page.
                try:
                    await page.close()
                    info_log("Page closed successfully.")
                except Exception as e_page_close:
                    error_log(f"Error closing page: {e_page_close}")
            else:
                info_log("Page was not initialized or already closed.")

            if 'browser' in locals() and browser and browser.is_connected():
                info_log("Closing browser.")
                try:
                    await browser.close()
                    info_log("Browser closed successfully.")
                except Exception as e_browser_close:
                    error_log(f"Error closing browser: {e_browser_close}")
            else:
                info_log("Browser was not initialized or already disconnected.")

# this code will run the main function
if __name__ == "__main__":
    # import asyncio # asyncio already imported at the top
    asyncio.run(main())